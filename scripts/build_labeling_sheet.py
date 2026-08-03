"""Build the hand-labeling spreadsheet for aspect ground truth (Hito 0 step 6).

The `aspect` field has no free proxy (unlike sentiment, which has the
star rating) -- see CLAUDE.md. This produces an .xlsx with a
deterministic, stratified 100-review sample (20 per star value) and
one dropdown-backed column per aspect, so labeling is point-and-click
rather than free typing.

Usage:
    python scripts/build_labeling_sheet.py
"""

from __future__ import annotations

import sys
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from aurapulse.data_loader import select_stratified_review_sample
from aurapulse.schemas import Aspect

REVIEW_SUBSET_PATH = Path("data/processed/review_subset.csv")
OUTPUT_PATH = Path("data/processed/aspect_labeling_sheet.xlsx")
PER_STAR = 20  # 5 star values * 20 = 100 reviews

ARIAL = "Arial"
YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
GREY_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(name=ARIAL, bold=True, color="FFFFFF")

# Aspect columns in a fixed, sensible order (matches the schema's enum order).
ASPECT_COLUMNS = [a.value for a in Aspect if a != Aspect.OTHER] + [Aspect.OTHER.value]


def build_instructions_sheet(wb: Workbook) -> None:
    """First tab: what this is, how to fill it, one example row."""
    # A freshly-constructed Workbook always has exactly one sheet, but
    # openpyxl's stubs type wb.active as possibly-None - index instead.
    ws = wb.worksheets[0]
    ws.title = "Instructions"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 100

    lines = [
        ("AuraPulse - Aspect Labeling Task", True, 14),
        ("", False, 11),
        (
            (
                "Why: sentiment has a free ground-truth proxy (the star rating), but aspect "
                "(what the review is actually about) doesn't. This sample lets us check how "
                "well the model extracts aspects against real human judgment."
            ),
            False,
            11,
        ),
        ("", False, 11),
        ("How to fill in the 'Labeling' tab:", True, 12),
        (
            "1. Read the review text in column D.",
            False,
            11,
        ),
        (
            (
                "2. For each aspect column (food, service, price, cleanliness, wait_time, "
                "ambience, other) the review actually talks about, pick 'positive', "
                "'negative', or 'neutral' from the dropdown."
            ),
            False,
            11,
        ),
        (
            (
                "3. Leave an aspect column BLANK if the review doesn't mention it at all. "
                "Most reviews will only have 1-3 aspects filled in, not all 7 - that's expected."
            ),
            False,
            11,
        ),
        (
            (
                "4. Only use 'other' when the review is clearly about something that doesn't "
                "fit the other 6 categories (e.g. parking, kids' menu, loyalty program). If you "
                "fill 'other', also write a short note in the other_detail column (e.g. "
                "'parking')."
            ),
            False,
            11,
        ),
        (
            (
                "5. The yellow columns (E through L) are the ones to fill in. Everything else "
                "is there for context only - don't edit it."
            ),
            False,
            11,
        ),
        ("", False, 11),
        ("Example (illustrative, not a real row in the sheet):", True, 12),
    ]
    for i, (text, bold, size) in enumerate(lines, start=1):
        cell = ws.cell(row=i, column=1, value=text)
        cell.font = Font(name=ARIAL, bold=bold, size=size)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        if text:
            ws.row_dimensions[i].height = max(18, 15 * (len(text) // 95 + 1))

    example_header_row = len(lines) + 1
    example_row = example_header_row + 1
    example_headers = ["text", "food", "service", "wait_time", "other", "other_detail"]
    example_values = [
        (
            "Great pasta but we waited 40 minutes to be seated, and the lot was full so we "
            "had to park two blocks away."
        ),
        "positive",
        "",
        "negative",
        "negative",
        "parking",
    ]
    for col, (header, value) in enumerate(zip(example_headers, example_values), start=1):
        h_cell = ws.cell(row=example_header_row, column=col, value=header)
        h_cell.font = Font(name=ARIAL, bold=True)
        v_cell = ws.cell(row=example_row, column=col, value=value)
        v_cell.font = Font(name=ARIAL)
        v_cell.alignment = Alignment(wrap_text=True, vertical="top")

    progress_row = example_row + 2
    label_cell = ws.cell(row=progress_row, column=1, value="Reviews labeled so far:")
    label_cell.font = Font(name=ARIAL, bold=True)
    count_cell = ws.cell(row=progress_row, column=2, value="=SUM(Labeling!M2:M101)")
    count_cell.font = Font(name=ARIAL, bold=True)
    total_cell = ws.cell(row=progress_row, column=3, value="/ 100")
    total_cell.font = Font(name=ARIAL)


def build_labeling_sheet(wb: Workbook, sample: pd.DataFrame) -> None:
    """Second tab: the actual data to label, one row per review."""
    ws = wb.create_sheet("Labeling")

    headers = (
        ["review_id", "business_id", "stars", "text"]
        + ASPECT_COLUMNS
        + ["other_detail", "labeled (auto)"]
    )
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    ws.freeze_panes = "E2"
    ws.row_dimensions[1].height = 30

    widths = {"A": 22, "B": 22, "C": 8, "D": 70}
    for col_letter in ["E", "F", "G", "H", "I", "J", "K"]:
        widths[col_letter] = 11
    widths["L"] = 28
    widths["M"] = 16
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width

    aspect_col_start = 5  # column E
    aspect_col_end = aspect_col_start + len(ASPECT_COLUMNS) - 1  # column K
    other_detail_col = aspect_col_end + 1  # column L
    labeled_col = other_detail_col + 1  # column M

    last_row = len(sample) + 1
    dv = DataValidation(type="list", formula1='"positive,negative,neutral"', allow_blank=True)
    dv.error = "Choose positive, negative, or neutral (or leave blank if not mentioned)."
    dv.errorTitle = "Invalid entry"
    ws.add_data_validation(dv)
    dv.add(
        f"{get_column_letter(aspect_col_start)}2:{get_column_letter(aspect_col_end)}{last_row}"
    )

    for i, row in enumerate(sample.itertuples(index=False), start=2):
        ws.cell(row=i, column=1, value=row.review_id).font = Font(name=ARIAL)
        ws.cell(row=i, column=2, value=row.business_id).font = Font(name=ARIAL)
        ws.cell(row=i, column=3, value=row.stars).font = Font(name=ARIAL)
        text_cell = ws.cell(row=i, column=4, value=row.text)
        text_cell.font = Font(name=ARIAL)
        text_cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[i].height = 75

        for col in range(aspect_col_start, aspect_col_end + 1):
            cell = ws.cell(row=i, column=col)
            cell.font = Font(name=ARIAL)
            cell.fill = YELLOW_FILL

        detail_cell = ws.cell(row=i, column=other_detail_col)
        detail_cell.font = Font(name=ARIAL)
        detail_cell.fill = YELLOW_FILL
        detail_cell.alignment = Alignment(wrap_text=True, vertical="top")

        labeled_cell = ws.cell(
            row=i,
            column=labeled_col,
            value=(
                f"=IF(COUNTA({get_column_letter(aspect_col_start)}{i}:"
                f"{get_column_letter(aspect_col_end)}{i})>0,1,0)"
            ),
        )
        labeled_cell.font = Font(name=ARIAL)
        labeled_cell.fill = GREY_FILL


def main() -> int:
    if not REVIEW_SUBSET_PATH.exists():
        print(f"ERROR: {REVIEW_SUBSET_PATH} not found. Run scripts/build_subset.py first.")
        return 1

    reviews = pd.read_csv(REVIEW_SUBSET_PATH)
    sample = select_stratified_review_sample(reviews, PER_STAR)
    print(f"Sampled {len(sample)} reviews ({PER_STAR} per star value)")

    wb = Workbook()
    build_instructions_sheet(wb)
    build_labeling_sheet(wb, sample)
    wb.active = 0  # land on Instructions when opened

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_PATH)
    print(f"Wrote {OUTPUT_PATH}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
